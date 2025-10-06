# automatizacion_descarga_oracle.py
import os
import sys
import traceback
import re
import oci
from openpyxl import load_workbook

# Intentar cargar .env si existe (opcional)
try:
    from dotenv import load_dotenv
    load_dotenv()
    print("[INFO] .env cargado (si existe).")
except Exception:
    print("[INFO] python-dotenv no disponible o .env no encontrado. Se usan variables de entorno del sistema.")

# ============================
# 🔧 CONFIGURACIÓN (EDITABLE / vía env)
# ============================
BUCKET_NAME = os.getenv("OCI_BUCKET")  # debe existir. Ej: "ABCD"
BASE_FOLDER = os.getenv("BASE_FOLDER", r"D:\ACC_2025\Descarga_Archivos_ORACLE\Guatavita")
EXCEL_FILE = os.getenv("EXCEL_FILE", r"D:\ACC_2025\Descarga_Archivos_ORACLE\Excel\Prueba_Guatavita.xlsx")
BUCKET_PREFIX = os.getenv("BUCKET_PREFIX", "pqrsdf")  # carpeta raíz dentro del bucket (puede dejarse "")

OCI_PROFILE = os.getenv("OCI_PROFILE", "DEFAULT")
OCI_CONFIG_FILE = os.path.expanduser(os.getenv("OCI_CONFIG_FILE", "~/.oci/config"))

# Columnas del Excel (1-indexed)
COL_ID = int(os.getenv("COL_ID", 1))
COL_CARPETA = int(os.getenv("COL_CARPETA", 2))
COL_FILENAME = int(os.getenv("COL_FILENAME", 3))
COL_NOMBRE_COMPLETO = int(os.getenv("COL_NOMBRE_COMPLETO", 4))
COL_ESTADO = int(os.getenv("COL_ESTADO", 5))

# Guardar cada cuántas filas (1 = guardar cada fila)
SAVE_EVERY_N_ROWS = int(os.getenv("SAVE_EVERY_N_ROWS", 1))

# ============================
# Funciones auxiliares
# ============================
def sanitize_filename(name):
    """Quita caracteres inválidos en Windows y recorta espacios."""
    if name is None:
        return ""
    name = str(name).strip()
    # eliminar caracteres: \ / : * ? " < > | 
    name = re.sub(r'[\\\/:\*\?"<>|]', "_", name)
    return name

def ensure_dir(path):
    if not os.path.exists(path):
        os.makedirs(path)

def log_exc(prefix=""):
    exc = traceback.format_exc()
    print(prefix + exc)

# ============================
# Validaciones previas
# ============================
if not BUCKET_NAME:
    print("[ERROR] La variable de entorno OCI_BUCKET no está definida.")
    print("▶ Define temporalmente en PowerShell con:  $env:OCI_BUCKET='NOMBRE'")
    print("▶ O crea un archivo .env con: OCI_BUCKET=NOMBRE")
    sys.exit(1)

if not os.path.isfile(EXCEL_FILE):
    print("[ERROR] No existe el archivo Excel especificado: {}".format(EXCEL_FILE))
    sys.exit(1)

# ============================
# Inicio del proceso
# ============================
try:
    print("[INFO] Conectando a Oracle Cloud (usando config: {}).".format(OCI_CONFIG_FILE))
    config = oci.config.from_file(OCI_CONFIG_FILE, OCI_PROFILE)
    object_storage = oci.object_storage.ObjectStorageClient(config)

    # obtener namespace desde OCI
    namespace = object_storage.get_namespace().data
    print("[OK] Conectado a OCI. Namespace detectado: {}".format(namespace))

    # cargar Excel
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    print("[INFO] Excel cargado: {}, filas a procesar (sin encabezado): {}".format(EXCEL_FILE, max(0, ws.max_row - 1)))

    # crear carpeta base local
    ensure_dir(BASE_FOLDER)
    print("[INFO] Carpeta base local: {}".format(BASE_FOLDER))

    total = max(0, ws.max_row - 1)
    encontrados = 0
    no_encontrados = 0
    guardados = 0

    # Recorrer filas
    for row in range(2, ws.max_row + 1):
        # leer y sanitizar valores
        raw_id = ws.cell(row, COL_ID).value
        raw_carpeta = ws.cell(row, COL_CARPETA).value
        raw_filename = ws.cell(row, COL_FILENAME).value
        raw_nombre_completo = ws.cell(row, COL_NOMBRE_COMPLETO).value

        file_id = sanitize_filename(raw_id) or "sin_id_row{}".format(row)
        carpeta = sanitize_filename(raw_carpeta)
        filename = sanitize_filename(raw_filename)
        nombre_completo = sanitize_filename(raw_nombre_completo) or filename

        print("\n[INFO] Procesando fila {} / {} - id: {} - carpeta: {} - filename: {}".format(row - 1, total, file_id, carpeta, filename))

        target_dir = os.path.join(BASE_FOLDER, file_id)
        ensure_dir(target_dir)

        target_file = os.path.join(target_dir, nombre_completo)

        # Validación de carpeta vacía
        if not carpeta:
            print("[WARN] Carpeta vacía en la fila {}, se marca NO.".format(row))
            ws.cell(row, COL_ESTADO).value = "NO - Carpeta vacía"
            no_encontrados += 1
            if SAVE_EVERY_N_ROWS and ((row - 1) % SAVE_EVERY_N_ROWS == 0):
                try:
                    wb.save(EXCEL_FILE)
                except Exception as e_save:
                    backup_file = EXCEL_FILE.replace(".xlsx", "_backup.xlsx")
                    wb.save(backup_file)
                    print("[WARN] Excel en uso, guardado en backup:", backup_file)
            continue

        # Validación de filename vacío
        if not filename:
            print("[WARN] filename vacío en la fila {}, se marca NO.".format(row))
            ws.cell(row, COL_ESTADO).value = "NO - filename vacío"
            no_encontrados += 1
            if SAVE_EVERY_N_ROWS and ((row - 1) % SAVE_EVERY_N_ROWS == 0):
                try:
                    wb.save(EXCEL_FILE)
                except Exception as e_save:
                    backup_file = EXCEL_FILE.replace(".xlsx", "_backup.xlsx")
                    wb.save(backup_file)
                    print("[WARN] Excel en uso, guardado en backup:", backup_file)
            continue

        # construir prefijo de carpeta dentro del bucket (si BUCKET_PREFIX existe, se incluye)
        if BUCKET_PREFIX:
            base_prefix = "{}/{}".format(BUCKET_PREFIX.strip('/'), carpeta.strip('/'))
        else:
            base_prefix = carpeta.strip('/')
        folder_prefix = base_prefix.rstrip('/') + '/'

        try:
            # 1) comprobar si "entramos" a la carpeta (listar con prefijo)
            print("[INFO] Listando objetos en bucket '{}' con prefijo '{}' ...".format(BUCKET_NAME, folder_prefix))
            list_resp = object_storage.list_objects(namespace, BUCKET_NAME, prefix=folder_prefix, limit=1)
            objs = list_resp.data.objects if list_resp.data else []
            if not objs:
                # carpeta no existe (no hay objetos con ese prefijo)
                print("[WARN] Carpeta NO encontrada en bucket: '{}' (prefijo {}).".format(BUCKET_NAME, folder_prefix))
                ws.cell(row, COL_ESTADO).value = "NO - Carpeta no encontrada"
                no_encontrados += 1
                if SAVE_EVERY_N_ROWS and ((row - 1) % SAVE_EVERY_N_ROWS == 0):
                    wb.save(EXCEL_FILE)
                continue
            else:
                print("[OK] Carpeta encontrada (hay objetos bajo el prefijo).")

            # 2) construir nombre exacto del objeto esperado
            object_expected = folder_prefix + filename
            print("[INFO] Buscando objeto exacto: '{}'".format(object_expected))
            exact_list = object_storage.list_objects(namespace, BUCKET_NAME, prefix=object_expected, limit=5)
            exact_objs = exact_list.data.objects if exact_list.data else []
            found_object_name = None

            # Verificar coincidencia exacta
            for o in exact_objs:
                if getattr(o, "name", None) == object_expected:
                    found_object_name = object_expected
                    break

            # 3) fallback: listar toda la carpeta y buscar coincidencias por "endswith"
            if not found_object_name:
                print("[INFO] No encontrado exacto, buscando coincidencias parciales dentro de la carpeta...")
                page = object_storage.list_objects(namespace, BUCKET_NAME, prefix=folder_prefix)
                cand = []
                for o in page.data.objects:
                    oname = getattr(o, "name", "")
                    # coincidencia por final de nombre
                    if oname.endswith("/" + filename) or oname.endswith(filename):
                        cand.append(oname)
                if cand:
                    found_object_name = cand[0]  # tomar la primera coincidencia
                    print("[INFO] Archivo hallado por coincidencia: {}".format(found_object_name))
                else:
                    print("[WARN] No se encontró '{}' dentro de la carpeta '{}'".format(filename, folder_prefix))

            # 4) Si se encontró un objeto, descargarlo
            if found_object_name:
                print("[INFO] Descargando objeto: {}".format(found_object_name))
                try:
                    get_resp = object_storage.get_object(namespace, BUCKET_NAME, found_object_name)
                    with open(target_file, "wb") as fh:
                        fh.write(get_resp.data.content)
                    ws.cell(row, COL_ESTADO).value = "SI"
                    encontrados += 1
                    print("[OK] Descargado y guardado en: {}".format(target_file))
                except Exception as e_get:
                    print("[ERROR] Falló descarga del objeto: {}".format(str(e_get)))
                    log_exc("[TRACE] ")
                    ws.cell(row, COL_ESTADO).value = "NO - fallo descarga: {}".format(str(e_get))
                    no_encontrados += 1
            else:
                # no encontrado
                ws.cell(row, COL_ESTADO).value = "NO - archivo no encontrado"
                no_encontrados += 1

        except Exception as e:
            print("[ERROR] Error al procesar la fila {}: {}".format(row, str(e)))
            log_exc("[TRACE] ")
            ws.cell(row, COL_ESTADO).value = "NO - error general: {}".format(str(e))
            no_encontrados += 1

        # Guardar progreso periódicamente (o cada fila si SAVE_EVERY_N_ROWS==1)
        if SAVE_EVERY_N_ROWS and ((row - 1) % SAVE_EVERY_N_ROWS == 0):
            try:
                wb.save(EXCEL_FILE)
                guardados += 1
                print("[INFO] Guardado intermedio del Excel ({}).".format(EXCEL_FILE))
            except Exception as e_save:
                print("[WARN] No se pudo guardar el Excel ahora: {}".format(str(e_save)))

    # Guardar al final
    try:
        wb.save(EXCEL_FILE)
        print("[INFO] Resultados guardados en Excel: {}".format(EXCEL_FILE))
    except Exception:
        backup_file = EXCEL_FILE.replace(".xlsx", "_backup.xlsx")
        wb.save(backup_file)
        print("[WARN] No se pudo guardar el Excel final. Guardado en:", backup_file)
        log_exc()

    print("\n[RESUMEN] Total: {}, Encontrados: {}, No encontrados: {}, Guardados parciales: {}"
          .format(total, encontrados, no_encontrados, guardados))

except Exception as e_main:
    print("[ERROR] Fallo general del script: {}".format(str(e_main)))
    log_exc()
    sys.exit(1)
