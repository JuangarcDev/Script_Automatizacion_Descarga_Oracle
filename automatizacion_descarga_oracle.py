import os
import oci
from openpyxl import load_workbook

# ============================
# 🔧 CONFIGURACIÓN (EDITABLE)
# ============================
BUCKET_NAME = os.getenv("OCI_BUCKET")  # nombre del bucket
BASE_FOLDER = os.getenv("BASE_FOLDER", r"D:\ACC_2025\Descarga_Archivos_ORACLE\Guatavita")  # carpeta base donde guardar
EXCEL_FILE = os.getenv("EXCEL_FILE", r"D:\ACC_2025\Descarga_Archivos_ORACLE\Excel\Prueba_Guatavita.xlsx")
BUCKET_PREFIX = "pqrsdf"  # carpeta raíz dentro del bucket (puede dejarse "" si no aplica)

# Config de OCI (se recomienda ~/.oci/config con perfil DEFAULT)
OCI_PROFILE = os.getenv("OCI_PROFILE", "DEFAULT")
OCI_CONFIG_FILE = os.getenv("OCI_CONFIG_FILE", "~/.oci/config")

# Columnas del Excel (1-indexed)
COL_ID = 1
COL_CARPETA = 2
COL_FILENAME = 3
COL_NOMBRE_COMPLETO = 4
COL_ESTADO = 5

# ============================
# 🔑 INICIALIZACIÓN
# ============================
try:
    print("[INFO] Conectando a Oracle Cloud...")
    config = oci.config.from_file(OCI_CONFIG_FILE, OCI_PROFILE)
    object_storage = oci.object_storage.ObjectStorageClient(config)

    # 🔹 Obtener namespace directamente de OCI
    namespace = object_storage.get_namespace().data
    print("[OK] Conectado correctamente a OCI. Namespace:", namespace)

    # 📖 Cargar Excel
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    print("[INFO] Archivo Excel cargado:", EXCEL_FILE)

    # Crear carpeta base si no existe
    if not os.path.exists(BASE_FOLDER):
        os.makedirs(BASE_FOLDER)
        print("[INFO] Carpeta base creada:", BASE_FOLDER)

    total = ws.max_row - 1
    encontrados, no_encontrados = 0, 0

    # ============================
    # 🚀 PROCESAMIENTO DE ARCHIVOS
    # ============================
    for row in range(2, ws.max_row + 1):
        file_id = str(ws.cell(row, COL_ID).value)
        carpeta = str(ws.cell(row, COL_CARPETA).value)
        filename = str(ws.cell(row, COL_FILENAME).value)
        nombre_completo = str(ws.cell(row, COL_NOMBRE_COMPLETO).value)

        # Crear subcarpeta del ID
        target_dir = os.path.join(BASE_FOLDER, file_id)
        if not os.path.exists(target_dir):
            os.makedirs(target_dir)

        target_file = os.path.join(target_dir, nombre_completo)

        try:
            # Construir ruta dentro del bucket
            if BUCKET_PREFIX:
                object_name = "{}/{}/{}".format(BUCKET_PREFIX, carpeta, filename)
            else:
                object_name = "{}/{}".format(carpeta, filename)

            # Descargar objeto
            response = object_storage.get_object(namespace, BUCKET_NAME, object_name)
            with open(target_file, "wb") as f:
                f.write(response.data.content)

            ws.cell(row, COL_ESTADO).value = "SI"
            encontrados += 1
            print("[OK] ({}/{}) Descargado: {}".format(row - 1, total, object_name))

        except Exception as e:
            ws.cell(row, COL_ESTADO).value = "NO"
            no_encontrados += 1
            print("[WARN] ({}/{}) No encontrado: {} -> {}".format(row - 1, total, object_name, str(e)))

    # Guardar cambios en Excel
    wb.save(EXCEL_FILE)
    print("[INFO] Resultados guardados en Excel:", EXCEL_FILE)
    print("[RESUMEN] Total: {}, Encontrados: {}, No encontrados: {}".format(total, encontrados, no_encontrados))

except Exception as e:
    print("[ERROR] Fallo general del script:", str(e))

